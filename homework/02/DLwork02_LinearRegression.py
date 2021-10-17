import numpy as np
from numpy.linalg import inv
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# 从xlsx文件中读取数据
def datapre():
    wb = load_workbook(filename='Folds5x2_pp.xlsx')
    ws = wb['Sheet1']

    ALL_data = []
    for rows in ws.rows:
        ALL_data.append(rows)

    pro_data = []
    for i in range(ws.max_row):
        pro_data.append([])
        for j in range(ws.max_column):
            pro_data[i].append(ALL_data[i][j].value)

    # 删除标题行
    del pro_data[0]
    # 转换为np数组
    pro_data_1 = np.asarray([sublist for sublist in pro_data])

    return pro_data_1

# 线性回归模型
def linearregression(x, w, b):

    y = np.dot(x,w) + b

    print("y: ", y)
    return y

# 损失函数
def loss_function(y, t):
    n = t.shape[0]
    return (0.5/n) * np.square((t - y)).sum()

# 梯度下降
def gradient_decent(x, t, w, b, lr=0.1):
    n = t.shape[0]
    y = linearregression(x, w, b)

    # 多元回归的偏导写法
    dw = (1/n) * np.dot(x.T, (y - t))
    w = w - lr * dw
    # 使偏导为0时的w
    # w = np.dot(np.dot(inv(np.dot(x.T,x)), x.T),t)

    db = (1/n) * (y-t)
    b = b - lr * db

    # 记录loss值
    loss = loss_function(y, t)
    lost_list.append(loss)

    return w, b

# 模型训练
def iterate(x, t, w, b, iternum):

    for i in range(iternum):
        w, b = gradient_decent(x, t, w, b)



lost_list = []

w = 0.01 * np.random.randn(4,1)
b = np.zeros(1)
# 准备数据并归一化
data = datapre()
col_max = np.amax(data, axis=0)
data_norm = data / col_max

# 提取x与t
x = data_norm[:, :-1]
t = data_norm[:, [-1]]

# 进行训练
iterate(x, t, w, b, 30)
# loss变化曲线
x_axis = range(0,30)
print("loss: ",lost_list)
plt.title("LOSS Curve")
plt.plot(x_axis,lost_list)
plt.show()
